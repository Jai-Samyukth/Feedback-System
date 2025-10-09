from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify, send_file
from werkzeug.utils import secure_filename
import os
import logging
from app.models.database import get_db
from app.models.student import Student
from app.services.excel_service import process_student_excel, create_sample_excel
from app.services.mapping_service import (
    process_mapping_excel, create_sample_mapping_excel,
    bulk_add_staff, bulk_add_subjects
)
from config import UPLOAD_FOLDER, ALLOWED_EXTENSIONS, MAX_FILE_SIZE
from utils import normalize_regno

logger = logging.getLogger(__name__)

admin_bp = Blueprint('admin', __name__)

# Ensure upload folder exists
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    """Check if file has an allowed extension."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@admin_bp.route('/admin_login', methods=['GET', 'POST'])
def admin_login():
    if request.method == 'POST':
        password = request.form.get('password')
        if password == 'vsbec':
            return redirect(url_for('admin.admin_dashboard'))
        else:
            flash("Incorrect password.", "danger")
            return redirect(url_for('admin.admin_login'))
    return render_template('admin_login.html')

@admin_bp.route('/admin/dashboard')
def admin_dashboard():
    return render_template('admin_dashboard.html')

@admin_bp.route('/admin/students', methods=['GET'])
def admin_students():
    """Display the student management page."""
    client = get_db()
    
    try:
        # Get departments from departments table
        dept_result = client.table('departments').select('name').order('name').execute()
        departments = [row['name'] for row in dept_result.data]
        
        # Get semesters from semesters table
        sem_result = client.table('semesters').select('name').order('name').execute()
        semesters = [row['name'] for row in sem_result.data]
    except Exception as e:
        logger.error(f"Error loading student management data: {e}")
        departments = []
        semesters = []
    
    return render_template('admin_students.html',
                         departments=departments,
                         semesters=semesters)

@admin_bp.route('/admin/students/list', methods=['GET'])
def list_students():
    """Get list of students filtered by department and semester."""
    department = request.args.get('department', '').strip()
    semester = request.args.get('semester', '').strip()
    
    if not department or not semester:
        return jsonify({
            'success': False,
            'message': 'Department and semester are required'
        })
    
    try:
        students = Student.get_by_dept_sem(department, semester)
        return jsonify({
            'success': True,
            'students': students,
            'count': len(students)
        })
    except Exception as e:
        logger.error(f"Error listing students: {e}")
        return jsonify({
            'success': False,
            'message': f'Error fetching students: {str(e)}'
        })

@admin_bp.route('/admin/students/add', methods=['POST'])
def add_students():
    """Add students via registration number range."""
    try:
        department = request.form.get('department', '').strip()
        semester = request.form.get('semester', '').strip()
        start_reg = request.form.get('startReg', '').strip()
        end_reg = request.form.get('endReg', '').strip()
        
        # Validate inputs
        if not all([department, semester, start_reg, end_reg]):
            return jsonify({
                'success': False,
                'message': 'All fields are required'
            })
        
        # Convert to integers
        start_num = int(start_reg)
        end_num = int(end_reg)
        
        if start_num > end_num:
            return jsonify({
                'success': False,
                'message': 'Start registration number must be less than or equal to end number'
            })
        
        if (end_num - start_num + 1) > 600:
            return jsonify({
                'success': False,
                'message': 'The range should not exceed 600 students'
            })
        
        # Prepare student data
        students_data = []
        for reg_no in range(start_num, end_num + 1):
            students_data.append((str(reg_no), department, semester))
        
        # Add students
        added_count, duplicate_count, duplicates = Student.bulk_add(students_data)
        
        if added_count > 0:
            message = f"Successfully added {added_count} students."
            if duplicate_count > 0:
                message += f" {duplicate_count} duplicates were skipped."
            return jsonify({
                'success': True,
                'message': message,
                'added': added_count,
                'duplicates': duplicate_count
            })
        else:
            return jsonify({
                'success': False,
                'message': f'All {duplicate_count} students already exist'
            })
    
    except ValueError:
        return jsonify({
            'success': False,
            'message': 'Invalid registration number format'
        })
    except Exception as e:
        logger.error(f"Error adding students: {e}")
        return jsonify({
            'success': False,
            'message': f'Error adding students: {str(e)}'
        })

@admin_bp.route('/admin/students/upload', methods=['POST'])
def upload_students_excel():
    """Upload students via Excel file."""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No file uploaded'
            })
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': 'No file selected'
            })
        
        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'message': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'
            })
        
        # Check file size
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({
                'success': False,
                'message': f'File too large. Maximum size is {MAX_FILE_SIZE / (1024*1024):.0f}MB'
            })
        
        # Save file
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        # Process file
        success, message, stats = process_student_excel(filepath)
        
        # Clean up uploaded file
        try:
            os.remove(filepath)
        except:
            pass
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'stats': stats
            })
        else:
            return jsonify({
                'success': False,
                'message': message,
                'stats': stats
            })
    
    except Exception as e:
        logger.error(f"Error uploading Excel: {e}")
        return jsonify({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        })

@admin_bp.route('/admin/students/delete', methods=['POST'])
def delete_student():
    """Delete a student."""
    try:
        registerno = request.form.get('registerno', '').strip()
        department = request.form.get('department', '').strip()
        semester = request.form.get('semester', '').strip()
        
        if not all([registerno, department, semester]):
            return jsonify({
                'success': False,
                'message': 'All fields are required'
            })
        
        # Normalize registration number
        registerno = normalize_regno(registerno)
        
        # Delete student
        deleted = Student.delete(registerno, department, semester)
        
        if deleted:
            return jsonify({
                'success': True,
                'message': f'Student {registerno} deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Student not found'
            })
    
    except Exception as e:
        logger.error(f"Error deleting student: {e}")
        return jsonify({
            'success': False,
            'message': f'Error deleting student: {str(e)}'
        })

@admin_bp.route('/admin/students/download-sample')
def download_sample():
    """Download a sample Excel file."""
    try:
        sample_path = os.path.join(UPLOAD_FOLDER, 'sample_students.xlsx')
        create_sample_excel(sample_path)
        return send_file(sample_path, as_attachment=True, download_name='sample_students.xlsx')
    except Exception as e:
        logger.error(f"Error creating sample file: {e}")
        flash('Error creating sample file', 'danger')
        return redirect(url_for('admin.admin_students'))

@admin_bp.route('/admin/students/delete-multiple', methods=['POST'])
def delete_multiple_students():
    """Delete multiple students at once."""
    try:
        data = request.get_json()
        students = data.get('students', [])
        
        if not students:
            return jsonify({
                'success': False,
                'message': 'No students selected for deletion'
            })
        
        client = get_db()
        deleted_count = 0
        errors = []
        
        for student in students:
            try:
                registerno = normalize_regno(student.get('registerno', ''))
                department = student.get('department', '').strip()
                semester = student.get('semester', '').strip()
                
                # Delete student
                result = client.table('students')\
                    .delete()\
                    .eq('registerno', registerno)\
                    .eq('department', department)\
                    .eq('semester', semester)\
                    .execute()
                
                if result.data:
                    deleted_count += 1
            except Exception as e:
                errors.append(f"Error deleting {student.get('registerno')}: {str(e)}")
        
        if deleted_count > 0:
            message = f'Successfully deleted {deleted_count} students.'
            if errors:
                message += f' {len(errors)} errors occurred.'
            return jsonify({
                'success': True,
                'message': message,
                'deleted': deleted_count,
                'errors': errors
            })
        else:
            return jsonify({
                'success': False,
                'message': 'No students were deleted. ' + (' '.join(errors) if errors else '')
            })
    
    except Exception as e:
        logger.error(f"Error deleting multiple students: {e}")
        return jsonify({
            'success': False,
            'message': f'Error deleting students: {str(e)}'
        })

@admin_bp.route('/admin/download-department-names')
def download_department_names():
    """Download department names as CSV file."""
    try:
        import csv
        import io
        
        client = get_db()
        dept_result = client.table('departments').select('name').order('name').execute()
        departments = [row['name'] for row in dept_result.data]
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Department Names'])
        writer.writerow(['⚠️ IMPORTANT: Use these EXACT names when uploading Excel files'])
        writer.writerow([])
        
        for dept in departments:
            writer.writerow([dept])
        
        # Prepare response
        output.seek(0)
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=department_names.csv'}
        )
    except Exception as e:
        logger.error(f"Error downloading department names: {e}")
        flash('Error downloading department names', 'danger')
        return redirect(url_for('admin.admin_students'))

@admin_bp.route('/admin/download-semester-names')
def download_semester_names():
    """Download semester names as CSV file."""
    try:
        import csv
        import io
        
        client = get_db()
        sem_result = client.table('semesters').select('name').order('name').execute()
        semesters = [row['name'] for row in sem_result.data]
        
        # Create CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['Semester Names'])
        writer.writerow(['⚠️ IMPORTANT: Use these EXACT names when uploading Excel files'])
        writer.writerow([])
        
        for sem in semesters:
            writer.writerow([sem])
        
        # Prepare response
        output.seek(0)
        from flask import Response
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': 'attachment; filename=semester_names.csv'}
        )
    except Exception as e:
        logger.error(f"Error downloading semester names: {e}")
        flash('Error downloading semester names', 'danger')
        return redirect(url_for('admin.admin_students'))

@admin_bp.route('/admin/download-reference-file')
def download_reference_file():
    """Download a comprehensive reference file with all departments and semesters."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        client = get_db()
        
        # Get departments
        dept_result = client.table('departments').select('name').order('name').execute()
        departments = [row['name'] for row in dept_result.data]
        
        # Get semesters
        sem_result = client.table('semesters').select('name').order('name').execute()
        semesters = [row['name'] for row in sem_result.data]
        
        # Get staff
        staff_result = client.table('staff').select('name').order('name').execute()
        staffs = [row['name'] for row in staff_result.data]
        
        # Get subjects
        subj_result = client.table('subjects').select('name').order('name').execute()
        subjects = [row['name'] for row in subj_result.data]
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        dept_sheet = wb.create_sheet('Departments')
        sem_sheet = wb.create_sheet('Semesters')
        staff_sheet = wb.create_sheet('Staff')
        subj_sheet = wb.create_sheet('Subjects')
        
        # Style for headers
        header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        header_font = Font(bold=True, size=12, color='000000')
        
        # Warning style
        warning_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        warning_font = Font(bold=True, size=11, color='FFFFFF')
        
        # Helper function to populate sheet
        def populate_sheet(sheet, title, items):
            # Add warning
            sheet['A1'] = '⚠️ CRITICAL: Use EXACT names from below when uploading Excel files'
            sheet['A1'].fill = warning_fill
            sheet['A1'].font = warning_font
            sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
            sheet.merge_cells('A1:B1')
            sheet.row_dimensions[1].height = 25
            
            # Add header
            sheet['A3'] = title
            sheet['A3'].fill = header_fill
            sheet['A3'].font = header_font
            sheet['A3'].alignment = Alignment(horizontal='center')
            sheet.column_dimensions['A'].width = 50
            
            # Add items
            for idx, item in enumerate(items, start=4):
                sheet[f'A{idx}'] = item
        
        # Populate sheets
        populate_sheet(dept_sheet, 'Department Names', departments)
        populate_sheet(sem_sheet, 'Semester Names', semesters)
        populate_sheet(staff_sheet, 'Staff Names', staffs)
        populate_sheet(subj_sheet, 'Subject Names', subjects)
        
        # Save to file
        reference_path = os.path.join(UPLOAD_FOLDER, 'reference_names.xlsx')
        wb.save(reference_path)
        
        return send_file(reference_path, as_attachment=True, download_name='REFERENCE_Names.xlsx')
    except Exception as e:
        logger.error(f"Error creating reference file: {e}")
        flash('Error creating reference file', 'danger')
        return redirect(url_for('admin.admin_students'))

@admin_bp.route('/admin', methods=['GET', 'POST'])
def admin():
    """Admin mapping page."""
    client = get_db()
    
    try:
        # Get departments
        dept_result = client.table('departments').select('name').order('name').execute()
        departments = [row['name'] for row in dept_result.data]
        
        # Get semesters
        sem_result = client.table('semesters').select('name').order('name').execute()
        semesters = [row['name'] for row in sem_result.data]
        
        # Get staff
        staff_result = client.table('staff').select('name').order('name').execute()
        staffs = [row['name'] for row in staff_result.data]
        
        # Get subjects
        subj_result = client.table('subjects').select('name').order('name').execute()
        subjects = [row['name'] for row in subj_result.data]
    except Exception as e:
        logger.error(f"Error loading admin page data: {e}")
        departments = []
        semesters = []
        staffs = []
        subjects = []
    
    if request.method == 'POST':
        department = request.form.get('department')
        semester = request.form.get('semester')
        staff_list = request.form.getlist('staff')
        subject_list = request.form.getlist('subject')
        
        new_mappings = [
            {'department': department, 'semester': semester, 'staff': staff.strip(), 'subject': subject.strip()}
            for staff, subject in zip(staff_list, subject_list)
            if staff.strip() and subject.strip()
        ]
        
        if not new_mappings:
            flash("Please enter at least one valid staff–subject mapping.", "danger")
        else:
            try:
                # Delete existing mappings for this dept/semester
                client.table('admin_mappings')\
                    .delete()\
                    .eq('department', department)\
                    .eq('semester', semester)\
                    .execute()
                
                # Insert new mappings
                client.table('admin_mappings').insert(new_mappings).execute()
                
                flash("Mapping(s) saved successfully.", "success")
            except Exception as e:
                logger.error(f"Error saving mappings: {e}")
                flash(f"Error saving mappings: {str(e)}", "danger")
            
            return redirect(url_for('admin.admin'))
    
    return render_template('admin_mapping.html',
                         departments=departments,
                         semesters=semesters,
                         staffs=staffs,
                         subjects=subjects)

@admin_bp.route('/admin/add_staff', methods=['POST'])
def add_staff():
    """Add a new staff member."""
    try:
        staff_name = request.form.get('staff_name', '').strip()
        if not staff_name:
            return jsonify({
                'success': False,
                'message': 'Staff name cannot be empty'
            })
        
        client = get_db()
        
        # Check if staff already exists
        existing = client.table('staff').select('name').eq('name', staff_name).execute()
        
        if existing.data:
            return jsonify({
                'success': False,
                'message': 'Staff name already exists'
            })
        
        # Insert new staff
        client.table('staff').insert({'name': staff_name}).execute()
        
        return jsonify({
            'success': True,
            'message': f'Successfully added staff: {staff_name}',
            'staff_name': staff_name
        })
    
    except Exception as e:
        logger.error(f"Error adding staff: {e}")
        return jsonify({
            'success': False,
            'message': f'Error adding staff: {str(e)}'
        })

@admin_bp.route('/admin/add_subject', methods=['POST'])
def add_subject():
    """Add a new subject."""
    try:
        subject_name = request.form.get('subject_name', '').strip()
        if not subject_name:
            return jsonify({
                'success': False,
                'message': 'Subject name cannot be empty'
            })
        
        client = get_db()
        
        # Check if subject already exists
        existing = client.table('subjects').select('name').eq('name', subject_name).execute()
        
        if existing.data:
            return jsonify({
                'success': False,
                'message': 'Subject already exists'
            })
        
        # Insert new subject
        client.table('subjects').insert({'name': subject_name}).execute()
        
        return jsonify({
            'success': True,
            'message': f'Successfully added subject: {subject_name}',
            'subject_name': subject_name
        })
    
    except Exception as e:
        logger.error(f"Error adding subject: {e}")
        return jsonify({
            'success': False,
            'message': f'Error adding subject: {str(e)}'
        })

@admin_bp.route('/admin/get_lists', methods=['GET'])
def get_lists():
    """Get staff and subject lists."""
    try:
        client = get_db()
        
        # Get staff list
        staff_result = client.table('staff').select('name').order('name').execute()
        staffs = [row['name'] for row in staff_result.data]
        
        # Get subjects list
        subj_result = client.table('subjects').select('name').order('name').execute()
        subjects = [row['name'] for row in subj_result.data]
        
        return jsonify({
            'success': True,
            'staffs': staffs,
            'subjects': subjects
        })
    except Exception as e:
        logger.error(f"Error getting lists: {e}")
        return jsonify({
            'success': False,
            'message': f'Error fetching lists: {str(e)}'
        })

@admin_bp.route('/admin/mappings/view', methods=['GET'])
def view_mappings():
    """View all staff-subject mappings."""
    client = get_db()
    
    try:
        # Get departments for filtering
        dept_result = client.table('departments').select('name').order('name').execute()
        departments = [row['name'] for row in dept_result.data]
        
        # Get semesters for filtering
        sem_result = client.table('semesters').select('name').order('name').execute()
        semesters = [row['name'] for row in sem_result.data]
    except Exception as e:
        logger.error(f"Error loading mappings view data: {e}")
        departments = []
        semesters = []
    
    return render_template('admin_view_mappings.html',
                         departments=departments,
                         semesters=semesters)

@admin_bp.route('/admin/mappings/list', methods=['GET'])
def list_mappings():
    """Get list of mappings filtered by department and semester."""
    department = request.args.get('department', '').strip()
    semester = request.args.get('semester', '').strip()
    
    try:
        client = get_db()
        query = client.table('admin_mappings').select('id, department, semester, staff, subject')
        
        if department and semester:
            query = query.eq('department', department).eq('semester', semester)
            query = query.order('staff').order('subject')
        elif department:
            query = query.eq('department', department)
            query = query.order('semester').order('staff').order('subject')
        elif semester:
            query = query.eq('semester', semester)
            query = query.order('department').order('staff').order('subject')
        else:
            query = query.order('department').order('semester').order('staff').order('subject')
            query = query.limit(500)
        
        result = query.execute()
        mappings = result.data
        
        return jsonify({
            'success': True,
            'mappings': mappings,
            'count': len(mappings)
        })
    except Exception as e:
        logger.error(f"Error listing mappings: {e}")
        return jsonify({
            'success': False,
            'message': f'Error fetching mappings: {str(e)}'
        })

@admin_bp.route('/admin/mappings/delete', methods=['POST'])
def delete_mapping():
    """Delete a specific mapping."""
    try:
        mapping_id = request.form.get('mapping_id', '').strip()
        
        if not mapping_id:
            return jsonify({
                'success': False,
                'message': 'Mapping ID is required'
            })
        
        client = get_db()
        result = client.table('admin_mappings').delete().eq('id', mapping_id).execute()
        
        if result.data:
            return jsonify({
                'success': True,
                'message': 'Mapping deleted successfully'
            })
        else:
            return jsonify({
                'success': False,
                'message': 'Mapping not found'
            })
    
    except Exception as e:
        logger.error(f"Error deleting mapping: {e}")
        return jsonify({
            'success': False,
            'message': f'Error deleting mapping: {str(e)}'
        })

@admin_bp.route('/admin/mappings/delete-all', methods=['POST'])
def delete_all_mappings():
    """Delete all mappings for a department and semester."""
    try:
        department = request.form.get('department', '').strip()
        semester = request.form.get('semester', '').strip()
        
        if not department or not semester:
            return jsonify({
                'success': False,
                'message': 'Department and semester are required'
            })
        
        client = get_db()
        result = client.table('admin_mappings')\
            .delete()\
            .eq('department', department)\
            .eq('semester', semester)\
            .execute()
        
        deleted_count = len(result.data) if result.data else 0
        
        return jsonify({
            'success': True,
            'message': f'Deleted {deleted_count} mappings successfully',
            'count': deleted_count
        })
    
    except Exception as e:
        logger.error(f"Error deleting mappings: {e}")
        return jsonify({
            'success': False,
            'message': f'Error deleting mappings: {str(e)}'
        })

@admin_bp.route('/admin/mappings/upload', methods=['POST'])
def upload_mapping_excel():
    """Upload staff-subject mappings via Excel file."""
    try:
        if 'file' not in request.files:
            return jsonify({
                'success': False,
                'message': 'No file uploaded'
            })
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({
                'success': False,
                'message': 'No file selected'
            })
        
        if not allowed_file(file.filename):
            return jsonify({
                'success': False,
                'message': 'Invalid file type. Please upload an Excel file (.xlsx or .xls)'
            })
        
        file.seek(0, os.SEEK_END)
        file_size = file.tell()
        file.seek(0)
        
        if file_size > MAX_FILE_SIZE:
            return jsonify({
                'success': False,
                'message': f'File too large. Maximum size is {MAX_FILE_SIZE / (1024*1024):.0f}MB'
            })
        
        filename = secure_filename(file.filename)
        filepath = os.path.join(UPLOAD_FOLDER, filename)
        file.save(filepath)
        
        replace_existing = request.form.get('replace_existing', 'false').lower() == 'true'
        
        success, message, stats = process_mapping_excel(filepath, replace_existing)
        
        try:
            os.remove(filepath)
        except:
            pass
        
        if success:
            return jsonify({
                'success': True,
                'message': message,
                'stats': stats
            })
        else:
            return jsonify({
                'success': False,
                'message': message,
                'stats': stats
            })
    
    except Exception as e:
        logger.error(f"Error uploading mapping Excel: {e}")
        return jsonify({
            'success': False,
            'message': f'Error processing file: {str(e)}'
        })

@admin_bp.route('/admin/mappings/download-sample')
def download_mapping_sample():
    """Download a sample Excel file for mappings."""
    try:
        sample_path = os.path.join(UPLOAD_FOLDER, 'sample_mapping.xlsx')
        create_sample_mapping_excel(sample_path)
        return send_file(sample_path, as_attachment=True, download_name='sample_staff_mapping.xlsx')
    except Exception as e:
        logger.error(f"Error creating sample mapping file: {e}")
        flash('Error creating sample file', 'danger')
        return redirect(url_for('admin.admin'))

@admin_bp.route('/admin/bulk-add', methods=['GET', 'POST'])
def bulk_add():
    """Page for bulk adding staff and subjects."""
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add_staff':
            staff_text = request.form.get('staff_text', '').strip()
            staff_list = [s.strip() for s in staff_text.split('\n') if s.strip()]
            
            if not staff_list:
                flash('Please enter at least one staff name', 'danger')
            else:
                added, duplicates = bulk_add_staff(staff_list)
                if added > 0:
                    message = f'Successfully added {added} staff members.'
                    if duplicates > 0:
                        message += f' {duplicates} duplicates were skipped.'
                    flash(message, 'success')
                else:
                    flash(f'No new staff added. All {duplicates} were duplicates.', 'warning')
        
        elif action == 'add_subjects':
            subjects_text = request.form.get('subjects_text', '').strip()
            subjects_list = [s.strip() for s in subjects_text.split('\n') if s.strip()]
            
            if not subjects_list:
                flash('Please enter at least one subject name', 'danger')
            else:
                added, duplicates = bulk_add_subjects(subjects_list)
                if added > 0:
                    message = f'Successfully added {added} subjects.'
                    if duplicates > 0:
                        message += f' {duplicates} duplicates were skipped.'
                    flash(message, 'success')
                else:
                    flash(f'No new subjects added. All {duplicates} were duplicates.', 'warning')
        
        return redirect(url_for('admin.bulk_add'))
    
    return render_template('admin_bulk_add.html')
